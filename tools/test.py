# Copyright (c) OpenMMLab. All rights reserved.
import argparse
import os
import os.path as osp

from mmengine.config import Config, DictAction
from mmengine.registry import RUNNERS
from mmengine.runner import Runner

from mmdet3d.utils import replace_ceph_backend
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl import Workbook
import torch.multiprocessing as mp
from utils import constants
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl import Workbook
import time

# TODO: support fuse_conv_bn and format_only
def parse_args():
    parser = argparse.ArgumentParser(
        description='MMDet3D test (and eval) a model')
    parser.add_argument('config', help='test config file path')
    parser.add_argument('checkpoint', help='checkpoint file')
    parser.add_argument(
        '--work-dir',
        help='the directory to save the file containing evaluation metrics')
    parser.add_argument(
        '--ceph', action='store_true', help='Use ceph as data storage backend')
    parser.add_argument(
        '--show', action='store_true', help='show prediction results')
    parser.add_argument(
        '--show-dir',
        help='directory where painted images will be saved. '
        'If specified, it will be automatically saved '
        'to the work_dir/timestamp/show_dir')
    parser.add_argument(
        '--task',
        type=str,
        choices=[
            'mono_det', 'multi-view_det', 'lidar_det', 'lidar_seg',
            'multi-modality_det'
        ],
        help='Determine the visualization method depending on the task.')
    parser.add_argument(
        '--wait-time', type=float, default=2, help='the interval of show (s)')
    parser.add_argument(
        '--cfg-options',
        nargs='+',
        action=DictAction,
        help='override some settings in the used config, the key-value pair '
        'in xxx=yyy format will be merged into config file. If the value to '
        'be overwritten is a list, it should be like key="[a,b]" or key=a,b '
        'It also allows nested list/tuple values, e.g. key="[(a,b),(c,d)]" '
        'Note that the quotation marks are necessary and that no white space '
        'is allowed.')
    parser.add_argument(
        '--launcher',
        choices=['none', 'pytorch', 'slurm', 'mpi'],
        default='none',
        help='job launcher')
    parser.add_argument('--local_rank', type=int, default=0)

    parser.add_argument('--sample_method', type=str, default='none')
    parser.add_argument('--sample_rate', type=str, default='1.0')
    parser.add_argument('--model_name', type=str, default='model')
    parser.add_argument('--feature_model', type=str, default='none')
    parser.add_argument('--dataset', type=str, default='kitti')

    args = parser.parse_args()

    if args.sample_method != 'none':
        constants.sample_method = args.sample_method
        constants.sample_rate = args.sample_rate
        constants.model_name = args.model_name
        constants.feature_model = args.feature_model
        constants.dataset = args.dataset
    print('constants test 2 aaaaa:', constants.model_name)

    if 'LOCAL_RANK' not in os.environ:
        os.environ['LOCAL_RANK'] = str(args.local_rank)
    return args


def trigger_visualization_hook(cfg, args):
    default_hooks = cfg.default_hooks
    if 'visualization' in default_hooks:
        visualization_hook = default_hooks['visualization']
        # Turn on visualization
        visualization_hook['draw'] = True
        if args.show:
            visualization_hook['show'] = True
            visualization_hook['wait_time'] = args.wait_time
        if args.show_dir:
            visualization_hook['test_out_dir'] = args.show_dir
        visualization_hook['vis_task'] = args.task
    else:
        raise RuntimeError(
            'VisualizationHook must be included in default_hooks.'
            'refer to usage '
            '"visualization=dict(type=\'VisualizationHook\')"')

    return cfg


def main():
    args = parse_args()
        
    # load config
    cfg = Config.fromfile(args.config)

    # TODO: We will unify the ceph support approach with other OpenMMLab repos
    if args.ceph:
        cfg = replace_ceph_backend(cfg)

    cfg.launcher = args.launcher
    if args.cfg_options is not None:
        cfg.merge_from_dict(args.cfg_options)

    # work_dir is determined in this priority: CLI > segment in file > filename
    if args.work_dir is not None:
        # update configs according to CLI args if args.work_dir is not None
        cfg.work_dir = args.work_dir
    elif cfg.get('work_dir', None) is None:
        # use config filename as default work_dir if cfg.work_dir is None
        cfg.work_dir = osp.join('./work_dirs',
                                osp.splitext(osp.basename(args.config))[0])

    cfg.load_from = args.checkpoint

    if args.show or args.show_dir:
        cfg = trigger_visualization_hook(cfg, args)

    # build the runner from config
    if 'runner_type' not in cfg:
        # build the default runner
        runner = Runner.from_cfg(cfg)
    else:
        # build customized runner from the registry
        # if 'runner_type' is set in the cfg
        runner = RUNNERS.build(cfg)

    # start testing
    runner.test()

def write_to_file(time):
    if constants.dataset == 'nus':
        file_name = 'result_nus.xlsx'
    else:
        file_name = 'result.xlsx'
    wb = load_workbook(file_name)
    if constants.model_name in wb.sheetnames:
        sheet = wb[constants.model_name]

        row = sheet.max_row
        if constants.dataset == 'nus':
            sheet.cell(row=row, column=column_index_from_string('N')).value = time
        else:
            sheet.cell(row=row, column=column_index_from_string('L')).value = time
        wb.save(file_name)



if __name__ == '__main__':
    time_start = time.time()
    main()
    time_end = time.time()

    time_cost = (time_end - time_start)/60

    if constants.sample_method != 'none':
        write_to_file(time_cost)

